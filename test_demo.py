from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.action_chains import ActionChains
import pytest
from pathlib import Path
from datetime import date
import openpyxl
from constants import globalConstants as gc


class Test_DemoClass:
    #Her testten önce çağırılır
    def setup_method(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        self.driver.get(gc.URL)
        #Günün tarihini al bu tarih ile bir klasör var mı kontrol et yoksa oluştur
        self.folderPath = str(date.today())
        Path(self.folderPath).mkdir(exist_ok=True)



    #Her testen sonra çağrılır
    def teardown_method(self):
        self.driver.quit()
    #setup -> test_demoFunc -> teardown
    def test_demoFunc(self):
        # 3A Act Arrange Assert
        text = "Hello"
        assert text == "Hello"
    #setup -> test_demo2 -> teardown   
    def test_demo2(self):
        assert True

    def getData():
        #veriyi al
        excelFile = openpyxl.load_workbook("data/invalid_login.xlsx")
        selectedSheet = excelFile["Sheet1"]
        totalRows = selectedSheet.max_row
        data=[]
        for i in range(2, totalRows+1):
            username = selectedSheet.cell(i,1).value    
            password = selectedSheet.cell(i,2).value
            tupleData = (username,password)
            data.append(tupleData)
        return data
    
        # return [("1","1"),("kullaniciadim","sifrem"),("kodlamaio","123")]

    # @pytest.mark.skip
    @pytest.mark.parametrize("username,password",getData())
    def test_invalid_login(self,username,password):
        self.waitForElementVisible((By.ID,"user-name"))
        userId = self.driver.find_element(By.NAME,"user-name")
        userId.send_keys(username)
        self.waitForElementVisible((By.ID,"password"))
        passId = self.driver.find_element(By.NAME,"password")
        passId.send_keys(password)
        self.waitForElementVisible((By.ID,"login-button"))
        logBut = self.driver.find_element(By.NAME,"login-button")
        logBut.click()
        self.waitForElementVisible((By.XPATH,"//*[@id='login_button_container']/div/form/div[3]/h3"))

        errorMsg = self.driver.find_element(By.XPATH,"//*[@id='login_button_container']/div/form/div[3]/h3")
        ErrorText = errorMsg.text
        self.driver.save_screenshot(f"{self.folderPath}/test-invalid-login-{username}-{password}.png")
        assert ErrorText == gc.ERROR_TEXT

    def waitForElementVisible(self,locator,timeout=5):
        WebDriverWait(self.driver,timeout).until(ec.visibility_of_element_located(locator))
        